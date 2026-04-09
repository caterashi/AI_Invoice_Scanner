"""
excel_export.py
===============
Generisanje Excel fajla iz liste InvoiceData objekata.
"""

from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ai_extractor import FIELDS, InvoiceData


# ─────────────────────────────────────────────────────────────────────────────
# Zaglavlja kolona (za Excel header red)
# ─────────────────────────────────────────────────────────────────────────────

HEADERS = {
    "BROJFAKT":   "Broj fakture",
    "DATUMF":     "Datum fakture",
    "DATUMPF":    "Datum prijema",
    "NAZIVPP":    "Naziv dobavljača",
    "SJEDISTEPP": "Adresa dobavljača",
    "IDPDVPP":    "ID/JIB broj",
    "JIBPUPP":    "PDV/PIB broj",
    "IZNBEZPDV":  "Iznos bez PDV",
    "IZNPDV":     "Iznos PDV",
    "IZNSAPDV":   "Ukupno s PDV",
}

# Kolone koje se formatiraju kao broj
AMOUNT_COLS = {"IZNBEZPDV", "IZNPDV", "IZNSAPDV"}


# ─────────────────────────────────────────────────────────────────────────────
# Javne funkcije
# ─────────────────────────────────────────────────────────────────────────────

def invoices_to_bytes(invoices: list[InvoiceData]) -> bytes:
    """
    Konvertuje listu InvoiceData objekata u Excel fajl (bytes).
    Spreman za st.download_button().
    """
    buf = io.BytesIO()
    _write_excel(invoices, buf)
    return buf.getvalue()


def invoices_to_dataframe(invoices: list[InvoiceData]) -> pd.DataFrame:
    """Konvertuje listu InvoiceData objekata u pandas DataFrame."""
    rows = [inv.to_dict() for inv in invoices]
    df   = pd.DataFrame(rows, columns=FIELDS)
    df.rename(columns=HEADERS, inplace=True)
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Interni Excel writer
# ─────────────────────────────────────────────────────────────────────────────

def _write_excel(invoices: list[InvoiceData], buf: io.BytesIO) -> None:
    """Piše formatirani Excel u buffer."""

    # Pripremi podatke
    rows = []
    for inv in invoices:
        row = {}
        for field in FIELDS:
            val = getattr(inv, field, "")
            if field in AMOUNT_COLS and val:
                try:
                    val = float(val)
                except ValueError:
                    pass
            row[HEADERS[field]] = val
        rows.append(row)

    df = pd.DataFrame(rows)

    # Piši u Excel
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Fakture")
        wb = writer.book
        ws = writer.sheets["Fakture"]

        _format_sheet(ws, df)

    buf.seek(0)


def _format_sheet(ws, df: pd.DataFrame) -> None:
    """Primijeni formatiranje na worksheet."""

    # Boje
    HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # tamno plava
    AMOUNT_FILL   = PatternFill("solid", fgColor="EBF3FB")   # svijetlo plava
    ALT_ROW_FILL  = PatternFill("solid", fgColor="F7FBFF")   # gotovo bijela
    WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

    THIN_BORDER = Border(
        left=Side(style="thin", color="BDD7EE"),
        right=Side(style="thin", color="BDD7EE"),
        top=Side(style="thin", color="BDD7EE"),
        bottom=Side(style="thin", color="BDD7EE"),
    )

    amount_col_indices = {
        i + 1
        for i, col in enumerate(df.columns)
        if col in [HEADERS[f] for f in AMOUNT_COLS]
    }

    # Header red
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER

    ws.row_dimensions[1].height = 32

    # Podatkovni redovi
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        is_alt = (row_idx % 2 == 0)
        for cell in row:
            col_idx = cell.column
            is_amount = col_idx in amount_col_indices

            cell.border    = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="right" if is_amount else "left",
                vertical="center",
            )

            if is_amount:
                cell.fill        = AMOUNT_FILL
                cell.number_format = '#,##0.00" KM"'
            else:
                cell.fill = ALT_ROW_FILL if is_alt else WHITE_FILL

        ws.row_dimensions[row_idx].height = 18

    # Širina kolona
    col_widths = {
        "Broj fakture":    16,
        "Datum fakture":   14,
        "Datum prijema":   14,
        "Naziv dobavljača": 35,
        "Adresa dobavljača": 35,
        "ID/JIB broj":     18,
        "PDV/PIB broj":    16,
        "Iznos bez PDV":   15,
        "Iznos PDV":       13,
        "Ukupno s PDV":    15,
    }

    for i, col in enumerate(df.columns, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = col_widths.get(col, 14)

    # Zamrzni header red
    ws.freeze_panes = "A2"
